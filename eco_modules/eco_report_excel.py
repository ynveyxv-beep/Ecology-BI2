"""
Excel report generation. Uses only metrics functions — no standalone calculations.
"""
from __future__ import annotations

import io
from datetime import datetime

import openpyxl
import pandas as pd
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

from .eco_metrics import (
    get_by_category,
    get_by_mro,
    get_by_omsu,
    get_kpi,
    get_time_series,
    get_trash_analysis,
)

# ─── Style constants ─────────────────────────────────────────────────────────

_THIN = Side(style="thin", color="CCCCCC")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
_ALT_FILL    = PatternFill("solid", fgColor="D6E4F0")

_HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF",  size=10)
_DATA_FONT   = Font(name="Calibri",            color="000000",  size=10)
_TITLE_FONT  = Font(name="Calibri", bold=True, color="1F4E79",  size=12)
_META_FONT   = Font(name="Calibri",            color="595959",  size=10)


# ─── Helpers ─────────────────────────────────────────────────────────────────

def _write_table(ws, df: pd.DataFrame, start_row: int, pct_cols: list[str] | None = None) -> int:
    """Write df as a formatted table starting at start_row. Returns the next free row."""
    pct_cols = pct_cols or []

    # Header
    for c, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=c, value=col_name)
        cell.font      = _HEADER_FONT
        cell.fill      = _HEADER_FILL
        cell.border    = _BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[start_row].height = 28

    # Rows
    for r_offset, row in enumerate(df.itertuples(index=False), 1):
        r = start_row + r_offset
        alt = r_offset % 2 == 0
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c)
            col_name = df.columns[c - 1]
            if col_name in pct_cols and isinstance(val, (int, float)) and not pd.isna(val):
                cell.value         = val / 100.0
                cell.number_format = "0.0%"
            else:
                cell.value = val
            cell.font      = _DATA_FONT
            cell.border    = _BORDER
            cell.alignment = Alignment(vertical="center")
            if alt:
                cell.fill = _ALT_FILL

    # Auto column width
    for c, col_name in enumerate(df.columns, 1):
        max_w = max(len(col_name), 6)
        for r in range(start_row + 1, start_row + len(df) + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None:
                max_w = max(max_w, len(str(v)))
        ws.column_dimensions[get_column_letter(c)].width = min(max_w + 4, 50)

    return start_row + len(df) + 2


def _section_title(ws, row: int, text: str):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = _TITLE_FONT
    return row + 1


# ─── Public API ──────────────────────────────────────────────────────────────

def export_excel(df: pd.DataFrame, filters_info: dict) -> bytes:
    """
    Build a multi-sheet Excel report from the filtered DataFrame.
    Numbers are sourced exclusively from metrics.py.

    Parameters
    ----------
    df : pd.DataFrame
        Already-filtered clean DataFrame.
    filters_info : dict
        Key → value pairs describing the active filters (written to the summary sheet).

    Returns
    -------
    bytes
        Raw bytes of the .xlsx file.
    """
    kpi         = get_kpi(df)
    cat_df      = get_by_category(df)
    omsu_df     = get_by_omsu(df)
    mro_df      = get_by_mro(df)
    ts_df       = get_time_series(df, "D")
    trash_top, trash_back = get_trash_analysis(df)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Сводка ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("Сводка")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 28

    ws["A1"] = "Отчёт центра экомониторинга"
    ws["A1"].font = Font(name="Calibri", bold=True, size=16, color="1F4E79")
    ws.row_dimensions[1].height = 28

    ws["A2"] = f"Сформирован: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font = _META_FONT

    row = 4
    row = _section_title(ws, row, "Применённые фильтры")
    for k, v in filters_info.items():
        ws.cell(row=row, column=1, value=k).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=v).font = _DATA_FONT
        row += 1

    row += 1
    row = _section_title(ws, row, "Ключевые показатели")
    kpi_rows = [
        ("Всего обращений",                 kpi["total"]),
        ("Выполнено",                        kpi["done"]),
        ("В работе",                         kpi["in_progress"]),
        ("% выполнения",                     kpi["pct_done"]),
        ("Муниципальных образований (ОМСУ)", kpi["omsu_count"]),
        ("Категорий",                        kpi["category_count"]),
        ("Период данных",                    kpi["period"]),
    ]
    for label, val in kpi_rows:
        ws.cell(row=row, column=1, value=label).font = Font(name="Calibri", bold=True, size=10)
        ws.cell(row=row, column=2, value=val).font = _DATA_FONT
        row += 1

    # ── По категориям ─────────────────────────────────────────────────────
    ws = wb.create_sheet("По категориям")
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, "Распределение обращений по категориям")
    _write_table(ws, cat_df, start_row=3, pct_cols=["Доля %"])

    # ── По ОМСУ ───────────────────────────────────────────────────────────
    ws = wb.create_sheet("По ОМСУ")
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, "Распределение по муниципальным образованиям")
    _write_table(ws, omsu_df, start_row=3, pct_cols=["Доля %", "Бэклог %"])

    # ── По МРО ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("По МРО")
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, "Распределение по межрайонным отделам")
    _write_table(ws, mro_df, start_row=3, pct_cols=["Доля %"])

    # ── Динамика ──────────────────────────────────────────────────────────
    ws = wb.create_sheet("Динамика")
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, "Динамика обращений по дням")
    if len(ts_df) > 0:
        ts_export = ts_df.copy()
        ts_export["Дата"] = ts_export["Дата"].dt.strftime("%d.%m.%Y")
        _write_table(ws, ts_export, start_row=3)
    else:
        ws.cell(row=3, column=1, value="Нет данных с датой").font = _META_FONT

    # ── Мусор ─────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Мусор")
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, 'Аналитика: "Мусор, свалки, стоки"')
    next_row = _section_title(ws, 3, "Топ ОМСУ по числу обращений")
    if len(trash_top) > 0:
        next_row = _write_table(ws, trash_top, start_row=next_row, pct_cols=["Доля %", "Бэклог %"])
    else:
        ws.cell(row=next_row, column=1, value="Нет данных").font = _META_FONT
        next_row += 2

    next_row = _section_title(ws, next_row, "Топ ОМСУ по доле незакрытых (бэклог)")
    if len(trash_back) > 0:
        _write_table(ws, trash_back, start_row=next_row, pct_cols=["Доля %", "Бэклог %"])
    else:
        ws.cell(row=next_row, column=1, value="Нет данных").font = _META_FONT

    # ── Данные ────────────────────────────────────────────────────────────
    ws = wb.create_sheet("Данные")
    ws.sheet_view.showGridLines = False
    _section_title(ws, 1, "Отфильтрованный срез (исходные строки после очистки)")

    data_export = df.copy()
    if "date" in data_export.columns:
        data_export["date"] = data_export["date"].dt.strftime("%d.%m.%Y").where(
            data_export["date"].notna(), other=""
        )
    if "time" in data_export.columns:
        data_export["time"] = data_export["time"].apply(
            lambda x: x.strftime("%H:%M") if pd.notna(x) else ""
        )

    data_export = data_export.rename(columns={
        "npp":      "№пп",
        "date":     "Дата",
        "req_no":   "Номер заявки",
        "time":     "Время",
        "category": "Категория",
        "mro":      "МРО",
        "omsu":     "ОМСУ",
        "status":   "Статус",
    })
    _write_table(ws, data_export, start_row=3)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()